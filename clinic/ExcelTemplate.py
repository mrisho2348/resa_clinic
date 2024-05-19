# views.py
import openpyxl
from django.http import HttpResponse

from clinic.models import Medicine, Service

def download_excel_template(request):
    # Create a workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Service Template"

    # Get the field names from the Service model, excluding auto fields and specific fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in Service._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Set the response to be an Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=service_template.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def download_medicine_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Medicine Template"
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at','remain_quantity','total_buying_price']
    model_fields = [field.name for field in Medicine._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

   
    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=medicine_template.xlsx'
    workbook.save(response)

    return response
