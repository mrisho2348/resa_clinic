from django.shortcuts import render
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models.functions import ExtractMonth
from django.db.models import Count
from datetime import datetime
from clinic.forms import YearSelectionForm
from clinic.models import PathodologyRecord, RemoteCompany, RemoteLaboratoryOrder, RemotePatient, RemoteProcedure, RemoteReferral, RemoteService
from openpyxl.utils import get_column_letter
from openpyxl.drawing.text import Paragraph, ParagraphProperties

# def render_comprehensive_report(year):
#     # Create a new Excel workbook
#     wb = Workbook()

#     # Get all distinct company names
#     all_companies = RemoteCompany.objects.values_list('name', flat=True)

#     # Query the database to get patient counts grouped by company and month
#     patients_by_company = (
#         RemotePatient.objects.filter(created_at__year=year)
#         .values('company__name')
#         .annotate(month=ExtractMonth('created_at'))
#         .annotate(num_patients=Count('id'))
#     )

#     # Organize the data into a dictionary
#     company_reports = {company: [0] * 13 for company in all_companies}  # Add one more column for the total
#     month_totals = [0] * 12

#     for patient in patients_by_company:
#         company_name = patient['company__name']
#         month = patient['month']
#         num_patients = patient['num_patients']

#         if month is not None:
#             month_index = month - 1  # ExtractMonth returns month as an integer
#             company_reports[company_name][month_index] = num_patients
#             company_reports[company_name][-1] += num_patients  # Add to the total column
#             month_totals[month_index] += num_patients  # Add to the total row

#     # Add the data to the Excel workbook
#     sheet = wb.active
#     sheet.title = f'Company Wise Report {year}'

#     # Add title
#     title_font = Font(size=14, bold=True, color="000000")
#     title_cell = sheet.cell(row=1, column=1)
#     title_cell.value = f'Company Wise Report for the Year {year}'
#     title_cell.font = title_font
#     title_cell.alignment = Alignment(horizontal='center')

#     # Merge title cells
#     sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)  # Adjust end_column to accommodate the total column

#     # Add subtitle
#     subtitle_font = Font(size=12, italic=True, color="808080")
#     subtitle_cell = sheet.cell(row=2, column=1)
#     subtitle_cell.value = 'Patient Counts by Company and Month'
#     subtitle_cell.font = subtitle_font
#     subtitle_cell.alignment = Alignment(horizontal='center')

#     # Merge subtitle cells
#     sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)  # Adjust end_column

#     # Add blank row
#     blank_row_cell = sheet.cell(row=3, column=1)
#     blank_row_cell.value = ''
#     blank_row_cell.alignment = Alignment(horizontal='center')

#     # Add headers
#     headers = ['Company'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']  # Add 'Total' header
#     for col, header in enumerate(headers, start=1):
#         header_cell = sheet.cell(row=4, column=col, value=header)
#         header_cell.font = Font(bold=True)
#         header_cell.alignment = Alignment(horizontal='center')

#     # Add data
#     row = 5
#     for company, counts in company_reports.items():
#         company_cell = sheet.cell(row=row, column=1, value=company)
#         company_cell.alignment = Alignment(horizontal='center')
#         for col, count in enumerate(counts, start=2):
#             count_cell = sheet.cell(row=row, column=col, value=count)
#             count_cell.alignment = Alignment(horizontal='center')
#         row += 1
        
#     overall_total = sum(month_totals)
#     # Add total row
#     total_row_cell = sheet.cell(row=row, column=1, value='Total')
#     total_row_cell.font = Font(bold=True)
#     total_row_cell.alignment = Alignment(horizontal='center')
#     for col, total in enumerate(month_totals, start=2):
#         total_cell = sheet.cell(row=row, column=col, value=total)
#         total_cell.alignment = Alignment(horizontal='center')

#     # Add total column
#     total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
#     total_column_cell.font = Font(bold=True)
#     total_column_cell.alignment = Alignment(horizontal='center')
#     for row, (company, counts) in enumerate(company_reports.items(), start=5):
#         total_cell = sheet.cell(row=row, column=len(headers), value=counts[-1])
#         total_cell.alignment = Alignment(horizontal='center')

    
#     # Add overall total cell
#     overall_total_cell = sheet.cell(row=row+1, column=len(headers), value=overall_total)  # Adjusted to the last column
#     overall_total_cell.font = Font(bold=True)
#     overall_total_cell.alignment = Alignment(horizontal='center')
    
#     red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
#     for col in range(1, len(headers) + 1):
#         sheet.cell(row=row+1, column=col).fill = red_fill
#     for row in range(3, row + 1):
#         sheet.cell(row=row+1, column=len(headers)).fill = red_fill
#     # Autofit column width
#     for col in range(1, len(headers) + 2):  # Adjusted the range to include the overall total column
#         max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, row + 1))
#         adjusted_width = max_length + 2
#         sheet.column_dimensions[chr(64 + col)].width = adjusted_width

#     # Prepare response to return the Excel workbook
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename="company_wise_report_{year}.xlsx"'
#     wb.save(response)

#     return response


def fetch_pathology_reports(year):
    # Get all distinct Pathology Record names for the given year
    all_pathology_records = PathodologyRecord.objects.values_list('name', flat=True)

    # Query the database to get patient counts grouped by Pathology Record and month for the given year
    patients_by_pathology_record = (
        PathodologyRecord.objects.filter(remoteconsultationnotes__created_at__year=year)
        .annotate(month=ExtractMonth('remoteconsultationnotes__created_at'))
        .values('name', 'month')
        .annotate(num_patients=Count('remoteconsultationnotes__id'))
    )

    # Organize the data into a dictionary
    pathology_record_reports = {record: [0] * 13 for record in all_pathology_records}  # Add one more column for the total
    month_totals = [0] * 12

    for patient in patients_by_pathology_record:
        pathology_record_name = patient['name']
        month = patient['month']
        num_patients = patient['num_patients']

        if month is not None:
            month_index = month - 1  # ExtractMonth returns month as an integer
            pathology_record_reports[pathology_record_name][month_index] = num_patients
            pathology_record_reports[pathology_record_name][-1] += num_patients  # Add to the total column
            month_totals[month_index] += num_patients  # Add to the total row

    return pathology_record_reports, month_totals

def render_pathology_report(sheet, year):
    # Fetch pathology report data for the given year
    pathology_reports, month_totals = fetch_pathology_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Pathology Record and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)  

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['CONSULTATIONS / PATHOLOGY'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']  
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for pathology_name, counts_by_month in pathology_reports.items():
        record_cell = sheet.cell(row=row, column=1, value=pathology_name)
        record_cell.alignment = Alignment(horizontal='left')  # Adjust alignment to left

        for col, count in enumerate(counts_by_month, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')

        row += 1

    # Add total row
    total_row = row  # Store the current row number for total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')
    
    # Add total column    
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row, (pathology_name, counts) in enumerate(pathology_reports.items(), start=5):
        total_cell = sheet.cell(row=row, column=len(headers), value=counts[-1])
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell
    overall_total = sum(month_totals)
    overall_total_cell = sheet.cell(row=row+1, column=len(headers), value=overall_total)  
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')
    
    # Apply red fill to the total column and total row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row+1, column=col).fill = red_fill
    for row in range(3, row + 1):
        sheet.cell(row=row+1, column=len(headers)).fill = red_fill

    # Autofit column width
    for col in range(1, len(headers) + 1):
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, total_row + 2))  # Adjusted range to include total row
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width

    # Return the modified sheet
    return sheet



def fetch_procedure_reports(year):
    # Get all services with the procedure category
    procedure_services = RemoteService.objects.filter(category='Procedure')

    # Query the database to get patient counts grouped by procedure category and month
    procedures_by_month = (
        RemoteProcedure.objects.filter(created_at__year=year)
        .annotate(month=ExtractMonth('created_at'))
        .values('name__name', 'month')
        .annotate(num_patients=Count('id'))
    )

    # Organize the data into a dictionary
    procedure_reports = {}
    for procedure_service in procedure_services:
        procedure_name = procedure_service.name
        procedure_reports[procedure_name] = [0] * 12  # Initialize counts for each month

    for procedure in procedures_by_month:
        procedure_name = procedure['name__name']
        month = procedure['month']
        num_patients = procedure['num_patients']

        if month is not None:
            month_index = int(month) - 1
            procedure_reports[procedure_name][month_index] = num_patients

    return procedure_reports


def render_procedure_reports(sheet, year):
    # Fetch procedure report data for the given year
    procedure_reports = fetch_procedure_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Procedure and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Procedure'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for procedure, counts in procedure_reports.items():
        procedure_cell = sheet.cell(row=row, column=1, value=procedure)
        procedure_cell.alignment = Alignment(horizontal='center')
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
        row += 1

    # Calculate overall totals for each month
    month_totals = [sum(counts) for counts in zip(*procedure_reports.values())]

    # Add total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add total column
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row_num, (procedure, counts) in enumerate(procedure_reports.items(), start=5):
        total_cell = sheet.cell(row=row_num, column=len(headers), value=sum(counts))
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell
    overall_total = sum(month_totals)
    overall_total_cell = sheet.cell(row=row, column=len(headers), value=overall_total)
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')

    # Apply red fill to the total column and total row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row, column=col).fill = red_fill
    for row in range(3, row + 1):
        sheet.cell(row=row, column=len(headers)).fill = red_fill

    # Autofit column width
    for col in range(1, len(headers) + 2):
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, row + 1))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width

    return sheet


  
  
def fetch_patient_laboratory_reports(year):
    # Get all distinct laboratory names
    laboratory_services = RemoteService.objects.filter(category='Laboratory')

    # Query the database to get patient counts grouped by laboratory and month
    laboratories_by_month = (
            RemoteLaboratoryOrder.objects.filter(created_at__year=year)
            .annotate(month=ExtractMonth('created_at'))
            .values('name__name', 'month')
            .annotate(num_patients=Count('id'))
        )

    # Organize the data into a dictionary
    laboratory_reports = {}
    for laboratory_service in laboratory_services:
        laboratory_name = laboratory_service.name
        laboratory_reports[laboratory_name] = [0] * 12 
    month_totals = [0] * 12

    for laboratory in laboratories_by_month:
        laboratory_name = laboratory['name__name']
        month = laboratory['month']
        num_patients = laboratory['num_patients']

        if month is not None:
                month_index = int(month) - 1
                laboratory_reports[laboratory_name][month_index] = num_patients

    return laboratory_reports, month_totals

def render_patient_laboratory_reports(sheet, year):
    # Fetch patient laboratory-wise report data for the given year
    laboratory_reports, month_totals = fetch_patient_laboratory_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Laboratory and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)  # Adjust end_column

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Laboratory'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']  # Add 'Total' header
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for laboratory, counts in laboratory_reports.items():
        laboratory_cell = sheet.cell(row=row, column=1, value=laboratory)
        laboratory_cell.alignment = Alignment(horizontal='center')
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
        row += 1
        
     # Calculate overall totals for each month
    month_totals = [sum(counts) for counts in zip(*laboratory_reports.values())]

    # Add total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add total column
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row_num, (laboratory, counts) in enumerate(laboratory_reports.items(), start=5):
        total_cell = sheet.cell(row=row_num, column=len(headers), value=sum(counts))
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell
    overall_total = sum(month_totals)
    overall_total_cell = sheet.cell(row=row, column=len(headers), value=overall_total)
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')

    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row, column=col).fill = red_fill
    for row in range(4, row + 1):
        sheet.cell(row=row, column=len(headers)).fill = red_fill
    # Autofit column width
    for col in range(1, len(headers) + 2):  # Adjusted the range to include the overall total column
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, row + 1))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width        
        
        
# Update render_comprehensive_report function to include pathology report
def render_comprehensive_report(year):
    # Create a new Excel workbook
    wb = Workbook()

    # Add company wise report
    company_wise_sheet = wb.active
    company_wise_sheet.title = f'Consult. Per Status Report {year}'
    render_patient_company_wise_reports(company_wise_sheet, year)

    # Add pathology report
    pathology_sheet = wb.create_sheet(title=f'Consult. Per Pathology {year}')
    render_pathology_report(pathology_sheet, year)

     # Add procedure report
    procedure_sheet = wb.create_sheet(title=f'Nursing Procedure Report {year}')
    render_procedure_reports(procedure_sheet, year)
    
     # Add laboratory result report
    laboratory_sheet = wb.create_sheet(title=f'Laboratory Tests Report {year}')
    render_patient_laboratory_reports(laboratory_sheet, year)
    
      # Add patient type report
    patient_type_sheet = wb.create_sheet(title=f'Patient Type Report {year}')
    render_patient_type_wise_reports(patient_type_sheet, year)
    
    # Add patient referral report
    referral_sheet = wb.create_sheet(title=f'Referral & MedEvac Report {year}')
    render_patient_referral_reports(referral_sheet, year)
    # Prepare response to return the Excel workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="comprehensive_report_{year}.xlsx"'
    wb.save(response)

    return response

def fetch_patient_company_wise_reports(year):
    # Get all distinct company names
    all_companies = RemoteCompany.objects.values_list('name', flat=True)

    # Query the database to get patient counts grouped by company and month
    patients_by_company = (
        RemotePatient.objects.filter(created_at__year=year)
        .values('company__name')
        .annotate(month=ExtractMonth('created_at'))
        .annotate(num_patients=Count('id'))
    )

    # Organize the data into a dictionary
    company_reports = {company: [0] * 13 for company in all_companies}  # Add one more column for the total
    month_totals = [0] * 12

    for patient in patients_by_company:
        company_name = patient['company__name']
        month = patient['month']
        num_patients = patient['num_patients']

        if month is not None:
            month_index = month - 1  # ExtractMonth returns month as an integer
            company_reports[company_name][month_index] = num_patients
            company_reports[company_name][-1] += num_patients  # Add to the total column
            month_totals[month_index] += num_patients  # Add to the total row

    return company_reports, month_totals

def render_patient_company_wise_reports(sheet, year):
    # Fetch patient company-wise report data for the given year
    company_reports, month_totals = fetch_patient_company_wise_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Company and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)  # Adjust end_column

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Site POB:'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']  # Add 'Total' header
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for company, counts in company_reports.items():
        company_cell = sheet.cell(row=row, column=1, value=company)
        company_cell.alignment = Alignment(horizontal='center')
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
        row += 1
        
    overall_total = sum(month_totals)
    # Add total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add total column
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row, (company, counts) in enumerate(company_reports.items(), start=5):
        total_cell = sheet.cell(row=row, column=len(headers), value=counts[-1])
        total_cell.alignment = Alignment(horizontal='center')

    
    # Add overall total cell
    overall_total_cell = sheet.cell(row=row+1, column=len(headers), value=overall_total)  # Adjusted to the last column
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row+1, column=col).fill = red_fill
    for row in range(3, row + 1):
        sheet.cell(row=row+1, column=len(headers)).fill = red_fill
    # Autofit column width
    for col in range(1, len(headers) + 2):  # Adjusted the range to include the overall total column
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, row + 1))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width

     
def fetch_patient_type_wise_reports(year):
    # Define the list of all patient types
    all_patient_types = ['National Staff', 'International Staff', 'National Visitor', 'International Visitor', 'Unknown Status', 'Others']

    # Query the database to get patient counts grouped by patient type and month
    patients_by_type = (
        RemotePatient.objects.filter(created_at__year=year)
        .values('patient_type')
        .annotate(month=ExtractMonth('created_at'))
        .annotate(num_patients=Count('id'))
    )

    # Organize the data into a dictionary
    patient_type_reports = {}   
    for patient_type in all_patient_types:
        # Initialize counts for each month
        patient_type_reports[patient_type] = [0] * 12
    month_totals = [0] * 12

    for patient in patients_by_type:
        patient_type = patient['patient_type']
        month = patient['month']
        num_patients = patient['num_patients']

        if month is not None:
            month_index = month - 1  # ExtractMonth returns month as an integer
            if patient_type not in patient_type_reports:
                patient_type_reports[patient_type] = [0] * 13  # Add one more column for the total
            patient_type_reports[patient_type][month_index] = num_patients
            patient_type_reports[patient_type][-1] += num_patients  # Add to the total column
            month_totals[month_index] += num_patients  # Add to the total row

    return patient_type_reports, month_totals

def render_patient_type_wise_reports(sheet, year):
    # Fetch patient type-wise report data for the given year
    patient_type_reports, month_totals = fetch_patient_type_wise_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Type and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)  # Adjust end_column

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Patient Type'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']  # Add 'Total' header
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for patient_type, counts in patient_type_reports.items():
        patient_type_cell = sheet.cell(row=row, column=1, value=patient_type)
        patient_type_cell.alignment = Alignment(horizontal='center')
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
        row += 1
        
    overall_total = sum(month_totals)
    # Add total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add total column
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row, (patient_type, counts) in enumerate(patient_type_reports.items(), start=5):
        total_cell = sheet.cell(row=row, column=len(headers), value=counts[-1])
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell
    overall_total_cell = sheet.cell(row=row+1, column=len(headers), value=overall_total)  # Adjusted to the last column
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row+1, column=col).fill = red_fill
    for row in range(3, row + 1):
        sheet.cell(row=row+1, column=len(headers)).fill = red_fill
    # Autofit column width
    for col in range(1, len(headers) + 2):  # Adjusted the range to include the overall total column
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, row + 1))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width
        
        
def fetch_patient_referral_reports(year):
    # Fetch data for patient referral report
    referrals = RemoteReferral.objects.filter(created_at__year=year)
    return referrals

def render_patient_referral_reports(sheet, year):
    # Fetch patient referral report data for the given year
    referrals = fetch_patient_referral_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add headers
    headers = ['Date', 'Patient mrn', 'First Name', 'Family Name', 'Age', 'Sex', 'Nationality',
               'Company', 'Patient Category', 'Med Evac/Refererred', 'Referral Reason', 'Transport Mode',
               'Patient Destination', 'Diagnosis']

    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=2, column=col)  # Note: Starting from row 2
        header_cell.value = header
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    for idx, referral in enumerate(referrals, start=3):  # Note: Starting from row 3
        age = referral.patient.age if referral.patient.dob else ''
        if referral.patient.dob:
            now = datetime.now()
            age = now.year - referral.patient.dob.year - ((now.month, now.day) < (referral.patient.dob.month, referral.patient.dob.day))

        diagnosis_content = ''
        for diagnosis_record in referral.patient.remotepatientdiagnosisrecord_set.all():
            for diagnosis in diagnosis_record.final_diagnosis.all():
                diagnosis_content += f'{diagnosis}, '

        # Parse notes
        notes_content = referral.notes.replace('<ol>', '').replace('<li>', '- ').replace('</li>', '\n').replace('</ol>', '')
        # Append referral data to Excel sheet
        sheet.cell(row=idx, column=1).value = referral.created_at.strftime('%d/%m/%Y')
        sheet.cell(row=idx, column=2).value = referral.patient.mrn
        sheet.cell(row=idx, column=3).value = referral.patient.first_name
        sheet.cell(row=idx, column=4).value = referral.patient.last_name
        sheet.cell(row=idx, column=5).value = f'{age} years'
        sheet.cell(row=idx, column=6).value = referral.patient.gender
        sheet.cell(row=idx, column=7).value = referral.patient.nationality.name
        sheet.cell(row=idx, column=8).value = referral.patient.company.name
        sheet.cell(row=idx, column=9).value = referral.patient.patient_type
        sheet.cell(row=idx, column=10).value = referral.nature_of_referral
        sheet.cell(row=idx, column=11).value = notes_content
        sheet.cell(row=idx, column=12).value = referral.transport_model        
        sheet.cell(row=idx, column=13).value = referral.destination_location
        sheet.cell(row=idx, column=14).value = diagnosis_content.rstrip(', ')

    # Autofit column width
    for col in range(1, 15):
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(2, len(referrals) + 3))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width

    # Save the workbook with the report
    return sheet

        
        
def generate_comprehensive_report(request):
    if request.method == 'POST':
        form = YearSelectionForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data['year']
            response = render_comprehensive_report(year)
            return response
    else:
        form = YearSelectionForm()
    
    return render(request, 'kahama_template/generate_comprehensive_report.html', {'form': form})        